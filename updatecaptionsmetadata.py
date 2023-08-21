#! /usr/bin/python3

import re
import sys
import os
import openpyxl
import warnings

warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)


recyclebin = 's3://s3-fremantle-uk-or-1/fremantleuk/DMS UK/FAST_CHANNEL_EDITS/Burbank/z_ToDelete/'
sheet      = '1. Master Metadata'
hns        = []

START_ROW = 4
START_COL = 2

#get house numbers from the command line and return the list
def get_HouseNumbers():

	hnList = []

	for n in range(1,len(sys.argv)):

		capf = sys.argv[n]

		if os.path.isfile(capf):
			strList = capf.split('.')
			hn  = strList[0]
			m   = re.match("^BUZ_[A-Z0-9]+",hn)
			if m:
				hnList.append(hn)

	return hnList

#get inputfrom the user and return it
def get_input(question):

	print()
	data = input(question)
	return data

#get the column number given the column name i.e. 'Supplier.Source' Fremantle.HouseNumber'
def getColNumNum(xlf,sheet,colname):


	for c in range(START_COL,100):

		txt = str(ws.cell(row=START_ROW,column=c).value)
		if txt == colname:
			return c

	print('\n   ~~~Could not find Column:', colname, 'in Metadata Sheet~~~\n')

	wb.save(xlf)
	wb.close()

	sys.exit(1)

	return c

def getxldata(ws,hn,epc,hnc,capc):

	r = START_ROW + 2

	fname     = ''
	capprefix = ''
	counter = 0 

	while counter < 10:

		txt = str(ws.cell(row=r,column=hnc).value)

		if txt == hn:

			fname     = str(ws.cell(row=r,column=epc).value)
			capprefix = str(ws.cell(row=r,column=capc).value)

			return fname, capprefix

		elif txt == 'None':

			counter += 1

		r += 1

	return fname, capprefix

#given the file name, return the file name with the version incrimented by 1
def getversion(txt):

	newfname = ''

	m = re.search('v(\d+)_(\d{8}\.[a-zA-Z]+)$',txt)

	if m:
		n       = int(m.group(1)) + 1
		
		matched = m.group(0)
		before   = txt[:m.start()]
		newfname = before + 'v' + str(n) + '_' + m.group(2)
		#print(txt, ':', newfname)

	else:
		n        = 2
		p        = re.search('(_\d{8}\.[a-zA-Z]+)$',txt)
		matched  = p.group(0)
		before   = txt[:p.start()]
		newfname = before + '_v' + str(n) + matched
		#print(txt, ':', newfname)

	return newfname

#dtermin if caption is scc or srt
def getcaptiontype(capf):

	scc = capf + '.scc'
	srt = capf + '.srt'

	if os.path.isfile(scc):
		return 'scc'
	elif os.path.isfile(srt):
		return 'srt'
	
	return ''



def updatexlf(ws,hncol,hn,epcol,newepname,capcol,caption):

	r = START_ROW + 2
	c = 0

	for r in range(r,1000):

		txt = str(ws.cell(row=r,column=hncol).value)

		if txt == hn:
			c = 1
			break

	if c == 1:

		#print('ROW =',r,': NAME-COL:',epcol,': SCC-COL:',capcol)


		cell = ws.cell(row=r,column=epcol)
		cell.value = newepname

		cell = ws.cell(row=r,column=capcol)
		cell.value = caption



#get info from the user 
vidpth  = get_input('Give me your video s3 path: ')
cappth  = get_input('Give me your captn s3 path: ')
xlf     = get_input('Give me your xl file name : ')
hns     = get_HouseNumbers()

#open the Metadata xlf for read/write
try:
	wb = openpyxl.load_workbook(xlf)
except:
	print('\n','   ~~~Cannot open Metadata Sheet~~~\n')
	sys.exit(1)

#open the '1. Master Metadata' sheet for read/write
try:
	ws = wb[sheet]
except:
	print('\n','   ~~~Cannot open Metadata Sheet~~~\n')
	sys.exit(1)


#get the column numbers for (filename, housenumber, scc prefix)
epcol  = getColNumNum(xlf,ws,'Supplier.OriginalName')
hncol  = getColNumNum(xlf,ws,'Fremantle.HouseNumber')
capcol = getColNumNum(xlf,ws,'TWK.AncillaryName')

error = False

#iterate over the house numbers index
for i in range(0,len(hns)):

	#get a house number from the hns list
	hn = hns[i]


	#given the house number, return the 'episode name' and 'caption prefix'
	epname, prefix = getxldata(ws,hn,epcol,hncol,capcol)

	#if the episode name or the caption prefix is blank, skip it
	if epname == '' or prefix == '':
		print(hn,': SKIPPING')
	else:
				
		newepname   = getversion(epname)       #get the new name of mxf file with the version incrimented by 1
		capext      = getcaptiontype(hn)       #get the extension tyoe fo the caption file
		capfname    = prefix + '.' + capext    #get the name of the caption with the extension
		parts       = newepname.split('.')     #split the new mxf name by .
		newcapname  = parts[0] + '.' + capext  #create the new caption file name

		#print(hn,':',newepname,':',newcapname)
		#BUZ_LMAD03247 : LetsMakeADeal_s2012_e4074_20230227.mxf : LetsMakeADeal_s2012_e4074_v2_20230227.mxf : LetsMakeADeal_s2012_e4074_20230227.scc
		#BUZ_LMAD03248 : LetsMakeADeal_s2012_e4075_20230227.mxf : LetsMakeADeal_s2012_e4075_v2_20230227.mxf : LetsMakeADeal_s2012_e4075_20230227.scc

		#link the caption file
		lncmd = 'ln ' + hn + '.' + capext + ' ' + newcapname
		print(lncmd)
		statln = os.system(lncmd)

		#move the captions to the recycle bin
		capmvcmd = 'aws s3 mv "' + vidpth + capfname + '" "' + recyclebin + '"'
		print(capmvcmd)
		statcapmv = os.system(capmvcmd)

		#update the xlf with the new names
		updatexlf(ws,hncol,hn,epcol,newepname,capcol,parts[0])

		
		#rename video file on s3
		vidmvcmd = 'aws s3 mv "' + vidpth + epname + '" "' + vidpth + newepname + '"'
		print(vidmvcmd)
		statvidmv = os.system(vidmvcmd)

		#print(statcapmv, type(statcapmv))
		#print(statvidmv, type(statvidmv))

		print('')
		print('   link    status:',str(statln))
		print('   recycle status:',str(statcapmv))
		print('   episode status:',str(statvidmv))
		print('')

		if (statln != 0) or (statcapmv != 0) or (statvidmv !=0):
			error = True

wb.save(xlf)
wb.close()

print('')
if error:	
	print('   Check the statuses, something failed')	
else:
	print('   S-U-C-C-E-S-S all statues are 0')
print('')



